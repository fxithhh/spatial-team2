import base64
import openpyxl
from PIL import Image
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import io
from openpyxl_image_loader import SheetImageLoader


def process_excel_file(file):
    """
    Processes the Excel file uploaded via form data to extract tabular data and embedded images.
    Returns images in Base64 JPEG format without MIME.
    """
    try:
        # Load workbook from the uploaded file (file-like object)
        workbook = load_workbook(file, data_only=True)
    except InvalidFileException:
        raise ValueError("Invalid Excel file format")
    except Exception as e:
        raise ValueError(f"Error loading Excel file: {str(e)}")

    # Check if the sheet exists
    sheet_name = "Output"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")
    else:
        output_sheet = workbook["Output"]

    # Extract data and images from the workbook
    extracted_data, base64_image_ls = extract_rows_and_images(output_sheet)
    column_headers = [cell.value for cell in output_sheet[1]]  # Assumes headers are in the first row
    column_headers.remove("Image")

    tabular_data = []
    for idx, row in enumerate(extracted_data):
        if "row_data" in row:
            row_data = dict(zip(column_headers, row["row_data"]))
            tabular_data.append(row_data)

    # Return as a dictionary
    result = {
        "data": tabular_data,  # List of dictionaries representing the rows
        "images": base64_image_ls  # List of standalone image Base64 strings
    }

    return result


def extract_rows_and_images(sheet):
    """
    Extracts rows and images from the workbook and associates images with row metadata.
    Returns images in Base64 JPEG format without MIME.
    """
    data = []
    base64_image_ls = []

    # Extract rows from the active sheet
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        data.append({"row_data": list(row), "row_index": idx})

    # Load images using SheetImageLoader
    image_loader = SheetImageLoader(sheet)

    # Check all cells for images
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row)):
        for cell in row:
            cell_coord = cell.coordinate
            try:
                # Attempt to retrieve the image
                image = image_loader.get(cell_coord)
                if image:
                    # Convert the image to JPEG Base64 without MIME
                    img_buffer = io.BytesIO()  # Create an in-memory buffer
                    image = image.convert("RGB")  # Ensure it's in RGB format for JPEG
                    image.save(img_buffer, format="JPEG")  # Save the image in JPEG format
                    img_buffer.seek(0)  # Move the pointer to the start of the buffer
                    image_base64 = base64.b64encode(img_buffer.getvalue()).decode("utf-8")  # Encode in Base64

                    base64_image_ls.append(image_base64)

                    # Locate the row in `data` and overwrite the corresponding cell with Base64
                    for entry in data:
                        if entry['row_index'] == row_idx:  # Match the row index
                            column_idx = cell.column - 1  # Convert column letter to zero-based index
                            entry['row_data'].pop(column_idx)
                            break
            except ValueError:
                # If no image is found, skip
                pass

    # Remove rows with too many invalid values (more than 4)
    data = [
        entry
        for entry in data
        if sum(1 for value in entry['row_data'] if value in [0, None]) <= 4
    ]

    # Re-adjust the row indices after removal
    for idx, entry in enumerate(data):
        entry['row_index'] = idx  # Update the row_index to reflect the new position

    return data, base64_image_ls
